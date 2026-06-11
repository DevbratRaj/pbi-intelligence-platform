"""
4_Governance_Engine.py — Action-Driven Governance Engine
PBI Intelligence Platform

Real analysis of an uploaded .pbit / .pbix file (via pbit_extractor.extract_pbit_metadata).
Every issue includes Problem · Impact · Exact Fix · Effort · Confidence · Action button.
"""

import re
import sys
import io
import json
import zipfile
from pathlib import Path
import streamlit as st
import pandas as pd

# ─────────────────────────────────────────────────────────────────────────────
# Page config + CSS
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Governance Engine — PBI Intelligence Platform",
    page_icon="📐",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
[data-testid="stSidebar"] { background-color: #0d1b2a; }
[data-testid="stSidebar"] * { color: #e0e8f0; }
[data-testid="stSidebar"] h1 {
    color:#fff; font-size:1.4rem; font-weight:700;
    padding-bottom:.5rem; border-bottom:1px solid #1e3a5f;
}
.impact-chip {
    display:inline-block; background:#eff6ff; color:#1d4ed8; border-radius:999px;
    padding:3px 12px; font-size:.75rem; font-weight:600; margin-right:4px; margin-bottom:4px;
}
.priority-badge-Critical { background:#fef2f2;color:#991b1b;border:1px solid #fca5a5;border-radius:4px;padding:2px 9px;font-size:.72rem;font-weight:700; }
.priority-badge-High     { background:#fff7ed;color:#9a3412;border:1px solid #fdba74;border-radius:4px;padding:2px 9px;font-size:.72rem;font-weight:700; }
.priority-badge-Medium   { background:#fefce8;color:#854d0e;border:1px solid #fde047;border-radius:4px;padding:2px 9px;font-size:.72rem;font-weight:700; }
.priority-badge-Low      { background:#f0fdf4;color:#166534;border:1px solid #86efac;border-radius:4px;padding:2px 9px;font-size:.72rem;font-weight:700; }
</style>
""", unsafe_allow_html=True)

with st.sidebar:
    st.markdown("<h1>PBI Intelligence Platform</h1>", unsafe_allow_html=True)
    st.markdown("---")

# ─────────────────────────────────────────────────────────────────────────────
# Import pbit_extractor from project root
# ─────────────────────────────────────────────────────────────────────────────
_HERE = Path(__file__).parent.parent
if str(_HERE) not in sys.path:
    sys.path.insert(0, str(_HERE))
from pbit_extractor    import extract_pbit_metadata
from pbit_fixer        import apply_safe_fixes, generate_rename_script
from dependency_graph  import build_graph, impact_of_rename


# ─────────────────────────────────────────────────────────────────────────────
# Fix-staging helpers — accumulate user-clicked fixes per file
# ─────────────────────────────────────────────────────────────────────────────
def _staged() -> dict:
    """Return the staged-fixes dict for the current file (creates if missing)."""
    if "staged_fixes" not in st.session_state:
        st.session_state["staged_fixes"] = {
            "descriptions":    [],   # [{table,measure,description}]
            "display_folders": [],   # [{table,measure,folder}]
            "placements":      [],   # [{measure,current_table,target_table}]
            "renames":         [],   # [{table,current_name,new_name}]
        }
    return st.session_state["staged_fixes"]


def _stage(category: str, item: dict) -> None:
    """Add a fix to the staging area, deduping by category+key fields."""
    s = _staged()
    bucket = s[category]
    # Dedupe by table+measure (or measure+current_table for placements)
    key_fields = {
        "descriptions":    ("table", "measure"),
        "display_folders": ("table", "measure"),
        "placements":      ("measure", "current_table"),
        "renames":         ("table", "current_name"),
    }[category]
    for existing in bucket:
        if all(existing.get(k) == item.get(k) for k in key_fields):
            existing.update(item)   # update in place
            return
    bucket.append(item)


# ─────────────────────────────────────────────────────────────────────────────
# Report usage helpers
# ─────────────────────────────────────────────────────────────────────────────
_ALWAYS_HIDDEN_VISUAL_TYPES: frozenset[str] = frozenset({
    "shape", "basicShape", "image", "textbox", "button", "actionButton",
})

_DATA_VISUAL_TYPES: frozenset[str] = frozenset({
    "barChart", "clusteredBarChart", "stackedBarChart", "hundredPercentStackedBarChart",
    "columnChart", "clusteredColumnChart", "stackedColumnChart", "hundredPercentStackedColumnChart",
    "lineChart", "areaChart", "stackedAreaChart", "hundredPercentStackedAreaChart",
    "lineClusteredColumnComboChart", "lineStackedColumnComboChart",
    "ribbonChart", "waterfallChart", "funnelChart", "scatterChart", "bubbleChart",
    "pieChart", "donutChart", "treemap", "sunburstChart",
    "tableEx", "pivotTable", "card", "multiRowCard", "kpi", "gauge",
    "map", "filledMap", "azureMap", "shapeMap", "slicer",
    "decompositionTreeVisual", "keyInfluencers", "qnaVisual",
    "rdlVisual", "pythonVisual", "rVisual", "scriptVisual", "customVisual",
})


def _is_data_visual(visual_type: str, field_count: int) -> bool:
    vt = visual_type.lower() if visual_type else ""
    if visual_type in _ALWAYS_HIDDEN_VISUAL_TYPES or vt in {v.lower() for v in _ALWAYS_HIDDEN_VISUAL_TYPES}:
        return False
    if vt == "slicer" and field_count == 0:
        return False
    if field_count == 0 and vt not in {v.lower() for v in _DATA_VISUAL_TYPES}:
        return False
    return True


def _clean_query_ref(ref: str) -> str:
    return re.sub(r"'([^']+)'\[", r"\1[", ref.strip())


def _extract_visual_title(config: dict, container: dict) -> str:
    sv = config.get("singleVisual", {})
    title_items = sv.get("vcObjects", {}).get("title", [])
    if isinstance(title_items, list) and title_items:
        props = title_items[0].get("properties", {})
        text_node = props.get("text", {}).get("expr", {})
        literal = text_node.get("Literal", {}).get("Value", "")
        if literal:
            return literal.strip("'")
        rpi = text_node.get("ResourcePackageItem", {}).get("name", "")
        if rpi:
            return rpi
    return container.get("name", "")


def _fields_with_roles(sv: dict, from_map: dict | None = None) -> list[dict]:
    rows: list[dict] = []
    seen: set[str] = set()
    from_map = from_map or {}

    projections = sv.get("projections", {})
    if isinstance(projections, dict):
        for role, items in projections.items():
            if not isinstance(items, list):
                continue
            for item in items:
                qr = _clean_query_ref(item.get("queryRef", ""))
                if qr and qr not in seen:
                    seen.add(qr)
                    rows.append({"Field": qr, "Role": role})

    proto_q = sv.get("prototypeQuery", {})
    for sel in proto_q.get("Select", []):
        field = ""
        native = sel.get("NativeReferenceName", "")
        if native and "[" in native:
            field = _clean_query_ref(native)

        if not field:
            for kind in ("Measure", "Column", "HierarchyLevel"):
                node = sel.get(kind, {})
                if not isinstance(node, dict) or not node:
                    continue
                src_ref = node.get("Expression", {}).get("SourceRef", {})
                entity = src_ref.get("Entity", "") or from_map.get(src_ref.get("Source", ""), "")
                prop = node.get("Property", "")
                if prop:
                    field = f"{entity}[{prop}]" if entity else f"[{prop}]"
                    break

        if not field:
            expr = sel.get("expression", sel.get("Expression", {}))
            if isinstance(expr, dict):
                for kind in ("Measure", "Column", "HierarchyLevel"):
                    node = expr.get(kind, {})
                    if not isinstance(node, dict) or not node:
                        continue
                    src_ref = node.get("Expression", {}).get("SourceRef", {})
                    entity = src_ref.get("Entity", "") or from_map.get(src_ref.get("Source", ""), "")
                    prop = node.get("Property", "")
                    if prop:
                        field = f"{entity}[{prop}]" if entity else f"[{prop}]"
                        break

        if field and field not in seen:
            seen.add(field)
            rows.append({"Field": field, "Role": "query"})

    return rows


def _parse_report_pages(raw_bytes: bytes) -> list[dict]:
    try:
        with zipfile.ZipFile(io.BytesIO(raw_bytes)) as zf:
            raw = zf.read("Report/Layout")
    except Exception:
        return []

    try:
        layout = json.loads(raw.decode("utf-16-le"))
    except Exception:
        return []

    pages_out = []
    for section in layout.get("sections", []):
        page_name = section.get("displayName") or section.get("name", "Unknown Page")
        visuals_out = []
        for container in section.get("visualContainers", []):
            try:
                config = json.loads(container.get("config", "{}"))
            except Exception:
                config = {}

            single_visual = config.get("singleVisual", {})
            visual_type = single_visual.get("visualType") or "unknown"
            title = _extract_visual_title(config, container)

            proto_q = single_visual.get("prototypeQuery", {})
            from_map = {
                f.get("Name", ""): f.get("Entity", "")
                for f in proto_q.get("From", []) if f.get("Name") and f.get("Entity")
            }
            field_rows = _fields_with_roles(single_visual, from_map)
            visuals_out.append({"title": title or visual_type, "type": visual_type, "fields": field_rows})
        pages_out.append({"name": page_name, "visuals": visuals_out})
    return pages_out


def _build_report_usage(meta: dict, pages: list[dict], broken_refs: list[dict]) -> dict:
    measure_keys = {f"{m['table']}[{m['name']}]": m for m in meta.get("measures", [])}
    measure_names = {m["name"]: m for m in meta.get("measures", [])}
    column_keys = {f"{c['table']}[{c['name']}]": c for c in meta.get("columns", []) if c.get("table")}

    usage_details: dict[str, dict] = {}
    for page in pages:
        for visual in page.get("visuals", []):
            if not _is_data_visual(visual.get("type", ""), len(visual.get("fields", []))):
                continue
            visual_name = visual.get("title") or visual.get("type", "unknown")
            visual_ref = f"{page['name']}::{visual_name}"
            for row in visual.get("fields", []):
                key = _clean_query_ref(row.get("Field", ""))
                if not key:
                    continue
                entry = usage_details.setdefault(key, {"pages": set(), "visuals": set()})
                entry["pages"].add(page["name"])
                entry["visuals"].add(visual_ref)

    usage_rows = []
    report_missing_rows = []
    for key, info in sorted(usage_details.items()):
        model_obj = measure_keys.get(key) or column_keys.get(key)
        is_measure = key in measure_keys
        is_column = key in column_keys
        name = key.split("[", 1)[1].rstrip("]") if "[" in key else key
        table = key.split("[", 1)[0] if "[" in key else ""
        row = {
            "Object": key,
            "Type": "Measure" if is_measure else ("Column" if is_column else "Unknown"),
            "Exists In Model": "Yes" if model_obj else "No",
            "Pages": ", ".join(sorted(info["pages"])) or "-",
            "Page Count": len(info["pages"]),
            "Visual Count": len(info["visuals"]),
            "Table": table,
            "Name": name,
        }
        usage_rows.append(row)
        if not model_obj:
            report_missing_rows.append(row)

    broken_measure_rows = []
    for b in broken_refs:
        src = measure_names.get(b["measure"])
        src_key = f"{src['table']}[{src['name']}]" if src else f"[{b['measure']}]"
        usage = usage_details.get(src_key)
        if usage is None:
            usage = next((v for k, v in usage_details.items() if k.endswith(f"[{b['measure']}]") ), None)
        broken_measure_rows.append({
            "Broken Measure": src_key,
            "Measure Table": src["table"] if src else "",
            "Error Type": b["kind"].title(),
            "Broken Reference": b["missing"],
            "Message": b["message"],
            "Still In Model": "Yes" if src else "No",
            "Used In Report": "Yes" if usage else "No",
            "Pages": ", ".join(sorted(usage["pages"])) if usage else "-",
            "Visual Count": len(usage["visuals"]) if usage else 0,
        })

    usage_index = {k: len(v["visuals"]) for k, v in usage_details.items()}
    return {
        "report_pages": pages,
        "usage_index": usage_index,
        "usage_rows": usage_rows,
        "report_missing_rows": report_missing_rows,
        "broken_measure_rows": broken_measure_rows,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Naming rule helpers
# ─────────────────────────────────────────────────────────────────────────────
_PASCAL_RE = re.compile(r"^[A-Z][a-zA-Z0-9]*$")
_ABBREVS   = {"qty","amt","rtn","cur","pct","yr","mth","cnt","avg","tot","num","dtl","desc","calc"}

def _check_name(name: str):
    """Return (reason, suggestion) or (None, None) if compliant."""
    if " " in name:
        return "Contains spaces", "".join(w.capitalize() for w in name.split())
    if re.search(r"[^A-Za-z0-9_]", name):
        clean = re.sub(r"[^A-Za-z0-9_]", "", name)
        return "Special characters in name", "".join(w.capitalize() for w in clean.split("_") if w)
    if "_" in name:
        return "Uses underscores (snake_case)", "".join(w.capitalize() for w in name.split("_") if w)
    if name == name.upper() and len(name) > 2:
        return "ALL_CAPS name", name.capitalize()
    if not _PASCAL_RE.match(name):
        return "Does not follow PascalCase", (name[0].upper() + name[1:]) if name else name
    low = name.lower()
    for a in _ABBREVS:
        if low == a or low.endswith("_" + a) or low.endswith(a) and len(low) <= len(a) + 4:
            return f"Abbreviation detected: '{a}' — prefer full word", name
    return None, None


# ─────────────────────────────────────────────────────────────────────────────
# Real analysis engine — works on output of extract_pbit_metadata()
# ─────────────────────────────────────────────────────────────────────────────
def _run_analysis(meta: dict) -> dict:
    measures = meta["measures"]
    tables   = meta["tables"]
    rels     = meta["relationships"]

    data_tables = {t["name"] for t in tables if t["columns"]}

    # ── Build the dependency graph FIRST — used to filter false positives
    #     and to power the rename-impact + accuracy panels.
    graph = build_graph(meta)
    hidden_set     = graph["hidden"]        # set of measure names
    time_intel_set = graph["time_intel"]    # set of measure names

    # 1. Naming
    naming_issues = []
    for m in measures:
        if m["name"] in hidden_set:
            continue   # don't lecture about names of hidden helper measures
        reason, suggestion = _check_name(m["name"])
        if reason:
            sev = "High" if any(c in m["name"] for c in (" ", "%", "#", "@", "$")) else "Medium"
            # Rename-impact ground truth from the dependency graph
            impact = impact_of_rename(graph, m["name"])
            naming_issues.append({
                "measure":   m["name"], "table": m["table"],
                "reason":    reason,
                "suggestion": suggestion or m["name"],
                "severity":  sev, "fixed": False,
                "impact_direct":     len(impact["direct"]),
                "impact_transitive": len(impact["transitive"]),
                "impact_total":      impact["total"],
                "impact_examples":   (impact["direct"] + impact["transitive"])[:5],
            })
    naming_score = max(0, 100 - int(len(naming_issues) / max(len(measures), 1) * 100))

    # 2. Descriptions  (skip hidden measures — they're internal helpers)
    desc_issues = []
    for m in measures:
        if m["name"] in hidden_set:
            continue
        ai = (
            f"Calculates {m['name']} for the selected filter context. "
            f"Underlying expression: {m['expression'][:90]}{'…' if len(m['expression']) > 90 else ''}"
            if m["expression"] else
            f"Metric tracking {m['name']} across the {m['table']} table."
        )
        desc_issues.append({
            "name":      m["name"], "table": m["table"],
            "has_desc":  bool(m["description"]),
            "description": m["description"] or "",
            "ai_suggest": ai,
        })
    has_desc = sum(1 for d in desc_issues if d["has_desc"])
    desc_score = int(has_desc / max(len(desc_issues), 1) * 100) if desc_issues else 100

    # 3. Placement  (skip time-intelligence — they belong with the date dim)
    # Detect existing dedicated measure tables: tables where all columns are hidden
    # (the common _Measures pattern: one dummy hidden column + all real measures)
    existing_measure_tables = sorted({
        t["name"] for t in tables
        if not t.get("is_internal")
        and t.get("measures")
        and all(c.get("is_hidden", False) for c in t.get("columns", []))
    })
    default_target = existing_measure_tables[0] if existing_measure_tables else "_Measures"

    misplaced = []
    for m in measures:
        if m["table"] not in data_tables:
            continue
        if m["name"] in time_intel_set:
            continue
        if m["name"] in hidden_set:
            continue
        expr = m.get("expression", "") or ""
        expr_preview = (expr[:100] + "…") if len(expr) > 100 else expr
        dependents = len(graph["callers"].get(m["name"], []))
        # Suggest the most natural target table: existing measure table > domain match > default
        suggested = default_target
        if existing_measure_tables:
            # If multiple measure tables, pick the one whose name most resembles the source table
            if len(existing_measure_tables) == 1:
                suggested = existing_measure_tables[0]
            else:
                best = next(
                    (t for t in existing_measure_tables
                     if any(kw in t.lower() for kw in m["table"].lower().split())),
                    existing_measure_tables[0],
                )
                suggested = best
        misplaced.append({
            "name":             m["name"],
            "current_table":    m["table"],
            "suggested_table":  suggested,
            "fixed":            False,
            "expression":       expr_preview,
            "dependents":       dependents,
        })
    placement_score = max(0, 100 - int(len(misplaced) / max(len(measures), 1) * 100))

    # 4. Folders  (skip hidden)
    no_folder = [
        {"name": m["name"], "table": m["table"], "fixed": False}
        for m in measures
        if not m["display_folder"] and m["name"] not in hidden_set
    ]
    folder_score = max(0, 100 - int(len(no_folder) / max(len(measures), 1) * 100))

    # 5. Bidirectional relationships
    bidir = [r for r in rels if str(r.get("cross_filter", "")).lower() in ("both", "bidirectional")]

    # 6. NEW — Accuracy findings (high-confidence, ground-truth from DAX)
    broken_refs = graph["broken"]                        # 100% confidence: refs to missing things
    dax_smells  = graph["smells"]                        # medium confidence: perf/style
    orphans     = graph["orphans"]                       # medium: unused measures (could be future-use)
    accuracy_score = max(0, 100 - len(broken_refs) * 10) # each broken ref = -10 points

    # Security + Schema audits (functions defined later — valid in Python)
    sec_audit      = _run_security_audit(meta)
    sch_audit      = _run_schema_audit(meta)
    security_score = sec_audit["security_score"]
    schema_score   = sch_audit["schema_score"]

    overall = int((naming_score + desc_score + placement_score + folder_score
                   + accuracy_score + security_score + schema_score) / 7)

    return {
        "overall": overall,
        "naming_score": naming_score, "desc_score": desc_score,
        "placement_score": placement_score, "folder_score": folder_score,
        "accuracy_score": accuracy_score,
        "security_score": security_score, "schema_score": schema_score,
        "naming_issues": naming_issues, "desc_issues": desc_issues,
        "misplaced": misplaced, "no_folder": no_folder, "bidir_rels": bidir,
        "broken_refs": broken_refs, "dax_smells": dax_smells, "orphans": orphans,
        "security_audit": sec_audit, "schema_audit": sch_audit,
        "hidden_count":     len(hidden_set),
        "time_intel_count": len(time_intel_set),
        "measure_count": len(measures),
        "table_count": len([t for t in tables if not t["is_internal"]]),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Top-5 priority builder
# ─────────────────────────────────────────────────────────────────────────────
def _build_top5(a: dict) -> list[dict]:
    items = []
    # Broken DAX references → ground-truth bugs, always rank #1
    if a.get("broken_refs"):
        items.append({
            "rank": 1, "priority": "Critical",
            "title": f"Fix {len(a['broken_refs'])} broken DAX reference(s)",
            "impact": "Measures reference missing tables / columns / measures — they fail at runtime or return BLANK silently",
            "effort": "10–30 min", "saving": "Restores measure correctness",
            "confidence": 100,
        })
    if a["bidir_rels"]:
        items.append({
            "rank": len(items)+1, "priority": "Critical",
            "title": f"Remove {len(a['bidir_rels'])} bidirectional relationship(s)",
            "impact": "Causes incorrect filter propagation — reports show wrong totals",
            "effort": "15 min", "saving": "Eliminates data accuracy risk",
            "confidence": 97,
        })
    # RLS / Security findings — compliance issue, very high priority
    sec = a.get("security_audit", {})
    for f in sec.get("findings", [])[:1]:
        if f["severity"] in ("Critical", "High"):
            items.append({
                "rank": len(items)+1, "priority": f["severity"],
                "title": f["finding"],
                "impact": (f["detail"][:120] + "…") if len(f["detail"]) > 120 else f["detail"],
                "effort": "30–60 min", "saving": "Prevents unauthorised data exposure",
                "confidence": 99,
            })
    # Schema: M2M and calculated columns misuse
    sch = a.get("schema_audit", {})
    if sch.get("m2m_rels"):
        items.append({
            "rank": len(items)+1, "priority": "Critical",
            "title": f"Fix {len(sch['m2m_rels'])} many-to-many relationship(s)",
            "impact": "M:M joins cause ambiguous aggregation and incorrect report totals",
            "effort": "1–2 hrs", "saving": "Eliminates data accuracy risk",
            "confidence": 96,
        })
    if sch.get("metric_calc_cols"):
        items.append({
            "rank": len(items)+1, "priority": "High",
            "title": f"Convert {len(sch['metric_calc_cols'])} calculated column(s) to measures",
            "impact": "Aggregation-style calculated columns inflate model size and reduce query speed",
            "effort": "30 min", "saving": "Reduces model memory footprint",
            "confidence": 90,
        })
    if a["naming_issues"]:
        high = sum(1 for i in a["naming_issues"] if i["severity"] == "High")
        items.append({
            "rank": len(items)+1, "priority": "High",
            "title": f"Rename {len(a['naming_issues'])} measures to PascalCase ({high} High severity)",
            "impact": "Inconsistent names confuse report authors and slow down model navigation",
            "effort": "30 min", "saving": "Saves ~3 hrs/month per developer",
            "confidence": 94,
        })
    if a["misplaced"]:
        items.append({
            "rank": len(items)+1, "priority": "High",
            "title": f"Move {len(a['misplaced'])} measures out of fact tables → _Measures",
            "impact": "Measures inside fact tables get accidentally deleted with columns",
            "effort": "20 min", "saving": "Prevents model breakage during maintenance",
            "confidence": 91,
        })
    nd = sum(1 for d in a["desc_issues"] if not d["has_desc"])
    if nd:
        items.append({
            "rank": len(items)+1, "priority": "Medium",
            "title": f"Add descriptions to {nd} measures",
            "impact": "Business users cannot understand measure meaning without descriptions",
            "effort": "1–2 hrs", "saving": "Reduces onboarding time by ~40%",
            "confidence": 88,
        })
    if a["no_folder"]:
        items.append({
            "rank": len(items)+1, "priority": "Low",
            "title": f"Assign display folders to {len(a['no_folder'])} measures",
            "impact": "Flat measure list slows authors finding the right metric",
            "effort": "45 min", "saving": "Saves 5 min per developer per session",
            "confidence": 85,
        })
    return items[:5]


# ─────────────────────────────────────────────────────────────────────────────
# Security / RLS audit
# ─────────────────────────────────────────────────────────────────────────────
def _run_security_audit(meta: dict) -> dict:
    """Audit RLS roles and return structured findings + a 0-100 score."""
    roles       = meta.get("roles", [])
    tables      = meta.get("tables", [])
    data_tables = [t for t in tables
                   if t.get("columns") and not t.get("is_internal") and not t.get("is_hidden")]

    _SENSITIVE_KW = {"sales","finance","hr","salary","payroll","customer","employee",
                     "budget","pnl","revenue","cost","profit","account","order","transaction"}
    sensitive_tables = [t for t in data_tables
                        if any(kw in t["name"].lower() for kw in _SENSITIVE_KW)]

    findings = []

    if not roles:
        sev = "Critical" if sensitive_tables else "High"
        findings.append({
            "severity": sev,
            "finding":  "No Row-Level Security (RLS) roles defined",
            "detail":   (
                f"Model has {len(data_tables)} data table(s) with no RLS configured. "
                "Any user with workspace access sees ALL rows."
                + (f" Sensitive-looking tables: {', '.join(t['name'] for t in sensitive_tables[:5])}."
                   if sensitive_tables else "")
            ),
            "fix": (
                "Power BI Desktop → Modelling → Manage Roles. "
                "Create a role with a DAX filter on each table containing restricted data, "
                "e.g.  [Region] = USERPRINCIPALNAME(). Then assign users in Power BI Service → Dataset → Security."
            ),
        })
        security_score = 0 if sensitive_tables else 30
    else:
        security_score = 75   # base credit for having at least one role
        for role in roles:
            perms       = role.get("table_permissions", [])
            empty_perms = [tp for tp in perms if not (tp.get("filter_dax") or "").strip()]
            protected   = {tp["table"] for tp in perms if (tp.get("filter_dax") or "").strip()}
            unprotected = [t for t in sensitive_tables if t["name"] not in protected]

            if empty_perms:
                findings.append({
                    "severity": "High",
                    "finding":  f"Role '{role['name']}' — {len(empty_perms)} table permission(s) have empty DAX filters",
                    "detail":   (
                        f"Tables with no filter expression: "
                        f"{', '.join(tp['table'] for tp in empty_perms[:5])}. "
                        "An empty filter grants unrestricted row access."
                    ),
                    "fix": "Add a DAX filter such as [Region] = USERPRINCIPALNAME() for each table permission.",
                })
            if unprotected:
                findings.append({
                    "severity": "Medium",
                    "finding":  f"Role '{role['name']}' — {len(unprotected)} sensitive table(s) not covered by any filter",
                    "detail":   f"Unfiltered sensitive tables: {', '.join(t['name'] for t in unprotected[:5])}.",
                    "fix": "Add table-level DAX filters in the role definition for these tables.",
                })

        if not findings:
            security_score = 95

    return {
        "roles":            roles,
        "findings":         findings,
        "sensitive_tables": [t["name"] for t in sensitive_tables],
        "security_score":   security_score,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Schema health audit
# ─────────────────────────────────────────────────────────────────────────────
def _run_schema_audit(meta: dict) -> dict:
    """Audit star-schema design, table naming, and calculated-column misuse."""
    tables  = meta.get("tables", [])
    rels    = meta.get("relationships", [])
    columns = meta.get("columns", [])

    data_tables = [t for t in tables
                   if t.get("columns") and not t.get("is_internal") and not t.get("is_hidden")]

    # ── 1. Table naming (Fact / Dim / Bridge prefix convention) ──────────
    _KNOWN_PREFIXES = ("fact","dim","bridge","xref","ref","cal","date","lookup","map","_")
    table_naming_issues = []
    for t in data_tables:
        n = t["name"]
        if any(n.lower().startswith(p) for p in _KNOWN_PREFIXES):
            continue
        lower = n.lower()
        guess = ("Fact" if any(k in lower for k in ("sale","order","trans","event","log","invoice","ticket"))
                 else "Dim")
        table_naming_issues.append({
            "table":      n,
            "reason":     "Doesn't follow Fact/Dim naming convention",
            "suggestion": f"{guess}{n}",
        })

    # ── 2. Isolated tables (no relationships at all) ──────────────────────
    connected = set()
    for r in rels:
        connected.add(r.get("from_table", ""))
        connected.add(r.get("to_table", ""))
    isolated_tables = [
        t for t in data_tables
        if t["name"] not in connected and not t["name"].startswith("_")
    ]

    # ── 3. Inactive relationships ─────────────────────────────────────────
    inactive_rels = [r for r in rels if not r.get("is_active", True)]

    # ── 4. Many-to-many relationships ─────────────────────────────────────
    m2m_rels = [
        r for r in rels
        if str(r.get("from_cardinality", "")).lower() == "many"
        and str(r.get("to_cardinality", "")).lower() == "many"
    ]

    # ── 5. Calculated columns that look like aggregations ─────────────────
    _AGG_KW = ("sum(","count(","average(","countrows(","sumx(","averagex(",
               "divide(","calculate(","max(","min(","maxx(","minx(")
    calc_cols = [c for c in columns if c.get("type") == "calculated" and c.get("expression", "")]
    metric_calc_cols = [
        c for c in calc_cols
        if any(kw in (c.get("expression", "") or "").lower() for kw in _AGG_KW)
    ]

    # ── Score ─────────────────────────────────────────────────────────────
    schema_score = 100
    schema_score -= min(30, len(table_naming_issues) * 5)
    schema_score -= min(20, len(isolated_tables) * 8)
    schema_score -= min(10, len(inactive_rels) * 3)
    schema_score -= min(25, len(m2m_rels) * 15)
    schema_score -= min(15, len(metric_calc_cols) * 5)
    schema_score = max(0, schema_score)

    return {
        "table_naming_issues": table_naming_issues,
        "isolated_tables":     isolated_tables,
        "inactive_rels":       inactive_rels,
        "m2m_rels":            m2m_rels,
        "calc_cols":           calc_cols,
        "metric_calc_cols":    metric_calc_cols,
        "schema_score":        schema_score,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Excel governance report builder
# ─────────────────────────────────────────────────────────────────────────────
def _build_excel_report(a: dict) -> bytes:
    """Return bytes of an .xlsx workbook containing all governance findings."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb  = openpyxl.Workbook()
    hdr = Font(bold=True, color="FFFFFF")
    bg  = PatternFill(fill_type="solid", fgColor="1E3A5F")

    def _sheet(title: str, headers: list, rows: list):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = hdr
            cell.fill = bg
            cell.alignment = Alignment(horizontal="center")
        for r in rows:
            ws.append(r)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                14, max((len(str(c.value or "")) for c in col), default=0) + 4)
        return ws

    # Summary
    ws_s = wb.active
    ws_s.title = "Summary"
    dims = [
        ("Naming Convention",  a["naming_score"]),
        ("Descriptions",       a["desc_score"]),
        ("Measure Placement",  a["placement_score"]),
        ("Display Folders",    a["folder_score"]),
        ("DAX Accuracy",       a["accuracy_score"]),
        ("Security / RLS",     a["security_score"]),
        ("Schema Health",      a["schema_score"]),
        ("OVERALL",            a["overall"]),
    ]
    ws_s.append(["Dimension", "Score (%)", "Status"])
    for cell in ws_s[1]:
        cell.font = hdr; cell.fill = bg
    for dim, sc in dims:
        status = "Good" if sc >= 80 else ("Needs Attention" if sc >= 60 else "Critical")
        ws_s.append([dim, sc, status])

    _sheet("Naming Issues",
           ["Table", "Measure", "Reason", "Suggested Name", "Severity"],
           [[i["table"], i["measure"], i["reason"], i["suggestion"], i["severity"]]
            for i in a["naming_issues"]])

    _sheet("Missing Descriptions",
           ["Table", "Measure", "AI-Generated Suggestion"],
           [[d["table"], d["name"], d["ai_suggest"]]
            for d in a["desc_issues"] if not d["has_desc"]])

    _sheet("Misplaced Measures",
           ["Measure", "Current Table", "Target Table"],
           [[m["name"], m["current_table"], m["suggested_table"]]
            for m in a["misplaced"]])

    sec = a.get("security_audit", {})
    _sheet("Security Findings",
           ["Severity", "Finding", "Detail", "Recommended Fix"],
           [[f["severity"], f["finding"], f["detail"], f["fix"]]
            for f in sec.get("findings", [])])

    sch = a.get("schema_audit", {})
    schema_rows: list = []
    for i in sch.get("table_naming_issues", []):
        schema_rows.append(["Table Naming",    i["table"],  i["reason"],            f"Rename to {i['suggestion']}"])
    for t in sch.get("isolated_tables", []):
        schema_rows.append(["Isolated Table",  t["name"],   "No relationships",     "Connect to Fact/Dim or remove"])
    for r in sch.get("m2m_rels", []):
        schema_rows.append(["Many-to-Many",    f"{r['from_table']} ↔ {r['to_table']}", "M:M join", "Add Bridge table"])
    for c in sch.get("metric_calc_cols", []):
        schema_rows.append(["Calc Col→Measure",f"{c.get('table','')}.{c['name']}",
                            (c.get("expression","") or "")[:80], "Convert to measure"])
    _sheet("Schema Findings", ["Type", "Object", "Detail", "Fix"], schema_rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# UI helpers
# ─────────────────────────────────────────────────────────────────────────────
def _score_col(s):  return "#16a34a" if s >= 80 else ("#ea580c" if s >= 60 else "#dc2626")
def _score_lbl(s):  return "Good" if s >= 80 else ("Needs Attention" if s >= 60 else "Critical")
def _priority_badge(p): return f'<span class="priority-badge-{p}">{p}</span>'


def _humanise_name(name: str) -> str:
    """Convert a technical name like 'TotalCollected' or 'FINANCIAL' to readable text."""
    # Handle dot-separated names e.g. "Financial.CollectorGroup" → "Financial Collector Group"
    name = name.replace(".", " ")
    # ALL_CAPS words → Title Case
    name = re.sub(r"\b([A-Z]{2,})\b", lambda m: m.group(1).capitalize(), name)
    # camelCase split: "TotalCollected" → "Total Collected"
    name = re.sub(r"([a-z])([A-Z])", r"\1 \2", name)
    # Sequence like "KPIValue" → "KPI Value"
    name = re.sub(r"([A-Z]+)([A-Z][a-z])", r"\1 \2", name)
    # Replace underscores/hyphens
    name = re.sub(r"[_\-]+", " ", name)
    return re.sub(r"\s+", " ", name).strip()


def _make_plain_desc(name: str, table: str, meta: dict, expression: str = "") -> str:
    """Generate a plain-English, non-technical description from a measure/field name + expression."""
    # 1. Use existing model description first
    measure = next(
        (m for m in meta.get("measures", []) if m["name"].lower() == name.lower()), None
    )
    if measure and measure.get("description"):
        return measure["description"]

    human = _humanise_name(name)
    n = name.lower().replace(".", " ").replace("_", " ")
    expr_lower = (expression or "").lower()

    # Expression-based clues (highest priority when expression is available)
    if expression:
        if "totalmtd" in expr_lower or "totalytd" in expr_lower or "totalqtd" in expr_lower:
            period = "month" if "mtd" in expr_lower else ("year" if "ytd" in expr_lower else "quarter")
            return f"Running total of {human} from the start of the {period} up to the selected date."
        if "datesbetween" in expr_lower or "datesytd" in expr_lower:
            return f"Calculates {human} over a specific date range."
        if "divide(" in expr_lower:
            return f"Calculates {human} as a ratio or percentage — divides two values safely."
        if "calculate(" in expr_lower and "filter(" in expr_lower:
            return f"Calculates {human} by applying specific filter conditions."
        if "distinctcount(" in expr_lower:
            return f"Counts the number of unique values for {human}."
        if "countrows(" in expr_lower:
            return f"Counts the number of records for {human}."
        if "average(" in expr_lower or "averagex(" in expr_lower:
            return f"The average value of {human}."
        if "sum(" in expr_lower or "sumx(" in expr_lower:
            return f"The total amount of {human} added together."
        if "max(" in expr_lower or "maxx(" in expr_lower:
            return f"The highest recorded value of {human}."
        if "min(" in expr_lower or "minx(" in expr_lower:
            return f"The lowest recorded value of {human}."

    # Name-based patterns (only when name clearly implies a concept)
    if re.search(r"\byoy\b|year.?on.?year", n):
        return f"Shows how {human} compares to the same period last year."
    if re.search(r"\bytd\b|year.?to.?date", n):
        return f"Running total of {human} from the start of the year to today."
    if re.search(r"\bmtd\b|month.?to.?date|monthlyrunning|runningcoll", n):
        return f"Running total of {human} from the start of the month to today."
    if re.search(r"\bbudget\b|\bforecast\b|\btarget\b", n):
        return f"The planned or budgeted value for {human}."
    if re.search(r"\bvariance\b", n):
        return f"The difference between the actual and expected value of {human}."
    if re.search(r"\bgrowth\b|\bchange\b", n):
        return f"How much {human} has grown or changed over the selected period."
    if re.search(r"\bavg\b|\baverage\b", n):
        return f"The average value of {human}."
    if re.search(r"\bpct\b|\bpercent\b|\bratio\b|\brate\b|\bmargin\b", n):
        return f"The percentage or rate for {human}."
    if re.search(r"\bcount\b|\bcnt\b|\brows\b", n):
        return f"The total count of {human}."
    if re.search(r"\bmax\b|\bmaximum\b|\bhighest\b", n):
        return f"The highest value of {human}."
    if re.search(r"\bmin\b|\bminimum\b|\blowest\b", n):
        return f"The lowest value of {human}."
    if re.search(r"\btotal\b|\bsum\b|\bamount\b|\bcollected\b|\bcollection\b", n):
        return f"The total amount of {human} added together."
    if re.search(r"\bworkingday\b|\bworkday\b|\bworking.?day\b", n):
        return f"The working day number within the current month."
    if re.search(r"\bcommission\b", n):
        return f"The commission amount for {human}."
    if re.search(r"\bactive\b|\bcurrent\b", n):
        return f"The current active count or value of {human}."
    if re.search(r"\bdate\b|\bmonth\b|\bquarter\b|\bweek\b", n):
        return f"The date / time value used to filter or group data — {human}."

    # Generic fallback
    t = _humanise_name(table) if table else ""
    return f"Shows {human} from the {t} data." if t else f"Shows {human}."


def _build_action_queue(a: dict) -> pd.DataFrame:
    rows = [
        {"Priority": "Critical", "Area": "Broken Measures", "Issues": len(a.get("broken_measure_rows", [])), "Action": "Fix DAX errors in measures"},
        {"Priority": "Critical", "Area": "Broken Report Fields", "Issues": len(a.get("report_missing_rows", [])), "Action": "Repair visuals using missing fields"},
        {"Priority": "Critical", "Area": "Relationships", "Issues": len(a["bidir_rels"]), "Action": "Remove bidirectional filters"},
        {"Priority": "High", "Area": "Security", "Issues": len(a["security_audit"]["findings"]), "Action": "Define or tighten RLS"},
        {"Priority": "High", "Area": "Placement", "Issues": len(a["misplaced"]), "Action": "Move measures to _Measures"},
    ]
    df = pd.DataFrame([r for r in rows if r["Issues"] > 0])
    if df.empty:
        return pd.DataFrame(columns=["Priority", "Area", "Issues", "Action"])
    order = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    return df.sort_values(by=["Priority", "Issues"], key=lambda s: s.map(order).fillna(99) if s.name == "Priority" else -s)


# ─────────────────────────────────────────────────────────────────────────────
# Page header
# ─────────────────────────────────────────────────────────────────────────────
st.title("📐 Governance Engine")
st.markdown(
    "Focus on the model issues that break trust, security, and maintainability. "
    "The sections below are ordered to help you fix the highest-risk problems first.")
st.markdown("---")

# ─────────────────────────────────────────────────────────────────────────────
# File source — uploaded file from main page, file uploader here, or demo data
# ─────────────────────────────────────────────────────────────────────────────
raw_bytes: bytes | None = st.session_state.get("pbi_file_bytes")

if not raw_bytes:
    st.info("📂 No file loaded yet. Upload a **.pbit / .pbix** for live analysis, or click **Use Demo Data**.")
    c1, c2 = st.columns([2, 1])
    with c1:
        up = st.file_uploader("Upload Power BI file", type=["pbit", "pbix"], label_visibility="collapsed")
        if up:
            raw_bytes = up.read()
            st.session_state["pbi_file_bytes"] = raw_bytes
            st.session_state.pop("gov_use_demo", None)
            for k in list(st.session_state.keys()):
                if k.startswith("gov_analysis_"): del st.session_state[k]
            st.success(f"✅ {up.name} loaded ({round(len(raw_bytes)/1024,1)} KB)")
            st.rerun()
    with c2:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("▶ Use Demo Data", use_container_width=True):
            st.session_state["gov_use_demo"] = True
            st.rerun()

use_demo = st.session_state.get("gov_use_demo", False) and not raw_bytes

# ─────────────────────────────────────────────────────────────────────────────
# Demo data
# ─────────────────────────────────────────────────────────────────────────────
_DEMO_MEASURES = [
    {"name":"total_sales","table":"FactSales","expression":"SUM(FactSales[Amount])","description":"","display_folder":"","is_hidden":False},
    {"name":"avg revenue","table":"FactSales","expression":"AVERAGE(FactSales[Revenue])","description":"","display_folder":"","is_hidden":False},
    {"name":"COGS_Amount","table":"FactCosts","expression":"SUM(FactCosts[Cost])","description":"","display_folder":"","is_hidden":False},
    {"name":"margin%","table":"FactSales","expression":"DIVIDE([NetRevenue],[TotalSales])","description":"","display_folder":"","is_hidden":False},
    {"name":"rtn_qty","table":"FactReturns","expression":"COUNTROWS(FactReturns)","description":"Count of returns","display_folder":"","is_hidden":False},
    {"name":"TotalSales","table":"_Measures","expression":"SUM(FactSales[Amount])","description":"Total sales amount","display_folder":"Revenue","is_hidden":False},
    {"name":"NetRevenue","table":"FactSales","expression":"[TotalSales]-[Returns]","description":"","display_folder":"","is_hidden":False},
    {"name":"YOYGrowthPct","table":"FactSales","expression":"DIVIDE([TotalSales]-[LYSales],[LYSales])","description":"","display_folder":"","is_hidden":False},
    {"name":"BudgetVariance","table":"FactBudget","expression":"[Actual]-[Budget]","description":"","display_folder":"","is_hidden":False},
    {"name":"ActiveCustomers","table":"DimCustomer","expression":"DISTINCTCOUNT(FactSales[CustomerID])","description":"Customers with transactions","display_folder":"KPIs","is_hidden":False},
]
_DEMO_TABLES = [
    {"name":"FactSales","columns":[{"name":"Amount"}],"measures":[],"is_internal":False,"is_hidden":False,"description":"","partitions":[]},
    {"name":"FactCosts","columns":[{"name":"Cost"}],"measures":[],"is_internal":False,"is_hidden":False,"description":"","partitions":[]},
    {"name":"FactBudget","columns":[{"name":"Budget"}],"measures":[],"is_internal":False,"is_hidden":False,"description":"","partitions":[]},
    {"name":"FactReturns","columns":[{"name":"Qty"}],"measures":[],"is_internal":False,"is_hidden":False,"description":"","partitions":[]},
    {"name":"DimCustomer","columns":[{"name":"ID"}],"measures":[],"is_internal":False,"is_hidden":False,"description":"","partitions":[]},
    {"name":"_Measures","columns":[],"measures":[],"is_internal":False,"is_hidden":False,"description":"","partitions":[]},
]
_DEMO_RELS = [{"from_table":"FactSales","from_column":"CustomerID","to_table":"DimCustomer","to_column":"ID",
               "from_cardinality":"many","to_cardinality":"one","cross_filter":"both","is_active":True}]

# Flat list mirroring what pbit_extractor produces — needed by the dependency
# graph to validate column references.
_DEMO_COLUMNS = [
    {"table":"FactSales", "name":"Amount"},
    {"table":"FactSales", "name":"Revenue"},
    {"table":"FactSales", "name":"CustomerID"},
    {"table":"FactCosts", "name":"Cost"},
    {"table":"FactBudget","name":"Budget"},
    {"table":"FactReturns","name":"Qty"},
    {"table":"DimCustomer","name":"ID"},
    {"table":"FactSales", "name":"TotalRevenue", "type":"calculated",
     "expression":"SUM(FactSales[Amount])", "is_hidden":False, "data_type":"double"},
]

_DEMO_ROLES: list = []
_DEMO_PAGES = [{
    "name": "Overview",
    "visuals": [{
        "title": "Sales Overview",
        "type": "tableEx",
        "fields": [
            {"Field": "FactSales[CustomerID]", "Role": "values"},
            {"Field": "_Measures[TotalSales]", "Role": "values"},
            {"Field": "FactSales[UnknownCol]", "Role": "values"},
        ],
    }],
}]

# ─────────────────────────────────────────────────────────────────────────────
# Get metadata: real or demo
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner="Analysing your model…")
def _get_meta_cached(b: bytes) -> dict:
    import tempfile, os
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pbit") as tmp:
        tmp.write(b); tmp_path = tmp.name
    try:
        return extract_pbit_metadata(tmp_path)
    finally:
        os.unlink(tmp_path)

if use_demo:
    meta = {
        "measures":      _DEMO_MEASURES,
        "tables":        _DEMO_TABLES,
        "columns":       _DEMO_COLUMNS,
        "relationships": _DEMO_RELS,
        "roles":         _DEMO_ROLES,
    }
    file_label = "Demo Data"
elif raw_bytes:
    try:
        full_meta = _get_meta_cached(raw_bytes)
        meta = full_meta
        file_label = Path(full_meta["file"]).name
        for e in full_meta.get("errors", []):
            st.warning(f"⚠️ {e}")
    except Exception as exc:
        st.error(f"Could not parse file: {exc}")
        st.stop()
else:
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# Run / cache analysis (per-file fingerprint)
# ─────────────────────────────────────────────────────────────────────────────
_fp = f"gov_analysis_{file_label}_{len(meta.get('measures', []))}"
if _fp not in st.session_state:
    st.session_state[_fp] = _run_analysis(meta)
analysis = st.session_state[_fp]

report_pages = _DEMO_PAGES if use_demo else (_parse_report_pages(raw_bytes) if raw_bytes else [])
analysis.update(_build_report_usage(meta, report_pages, analysis["broken_refs"]))
st.session_state[_fp] = analysis

# ─────────────────────────────────────────────────────────────────────────────
# Source badge + Reset button
# ─────────────────────────────────────────────────────────────────────────────
b1, b2 = st.columns([4, 1])
with b1:
    st.markdown(
        f'<span style="background:#dbeafe;color:#1e40af;border-radius:4px;padding:3px 12px;'
        f'font-size:.82rem;font-weight:600;">📂 Source: {file_label}</span>'
        f'&nbsp;&nbsp;'
        f'<span style="background:#f3f4f6;color:#374151;border-radius:4px;padding:3px 12px;font-size:.82rem;">'
        f'{analysis["measure_count"]} measures · {analysis["table_count"]} tables</span>',
        unsafe_allow_html=True,
    )
with b2:
    if st.button("🔄 Reset / change file", use_container_width=True):
        st.session_state.pop("pbi_file_bytes", None)
        st.session_state.pop("gov_use_demo", None)
        for k in list(st.session_state.keys()):
            if k.startswith("gov_analysis_"): del st.session_state[k]
        st.rerun()

st.markdown("<br>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# COMPACT ACTION OVERVIEW
# ─────────────────────────────────────────────────────────────────────────────
overall = analysis["overall"]
broken_used = sum(1 for row in analysis["broken_measure_rows"] if row["Used In Report"] == "Yes")
report_missing = len(analysis["report_missing_rows"])
critical_issues = len(analysis["broken_measure_rows"]) + report_missing + len(analysis["bidir_rels"])

mc1, mc2, mc3 = st.columns(3)
with mc1:
    st.metric("Broken Measures", len(analysis["broken_measure_rows"]))
with mc2:
    st.metric("Broken Fields On Pages", report_missing)
with mc3:
    st.metric("Broken Measures Used", broken_used)

st.markdown("#### What Needs Attention")
queue_df = _build_action_queue(analysis)
if queue_df.empty:
    st.success("No material governance issues detected.")
else:
    st.dataframe(queue_df, use_container_width=True, hide_index=True)
    st.caption("Start with broken objects that are actually used on report pages. Cleanup tabs are kept separate.")

st.markdown("---")

# ─────────────────────────────────────────────────────────────────────────────
# DETAILED ACTION TABS
# ─────────────────────────────────────────────────────────────────────────────
tab_broken, tab_report, tab_place, tab_rel, tab_acc, tab_sec, tab_name, tab_desc, tab_fold, tab_schema = st.tabs([
    f"🚨 Broken Objects ({len(analysis['broken_measure_rows'])})",
    f"📑 Report Usage ({len(analysis['report_missing_rows'])})",
    f"📂 Placement ({len(analysis['misplaced'])})",
    f"🔗 Relationships ({len(analysis['bidir_rels'])})",
    f"🎯 Accuracy ({len(analysis['dax_smells']) + len(analysis['orphans'])})",
    f"🔒 Security ({len(analysis['security_audit']['findings'])})",
    f"Optional: Naming ({len(analysis['naming_issues'])})",
    f"Optional: Descriptions ({sum(1 for d in analysis['desc_issues'] if not d['has_desc'])})",
    f"Optional: Folders ({len(analysis['no_folder'])})",
    f"Optional: Schema ({len(analysis['schema_audit']['table_naming_issues']) + len(analysis['schema_audit']['isolated_tables']) + len(analysis['schema_audit']['m2m_rels']) + len(analysis['schema_audit']['metric_calc_cols'])})",
])

# ── TAB 1: Broken objects ────────────────────────────────────────────────────
with tab_broken:
    broken_rows = analysis["broken_measure_rows"]
    if not broken_rows:
        st.success("✅ No broken measures detected in DAX.")
    else:
        st.dataframe(
            pd.DataFrame(broken_rows).sort_values(by=["Used In Report", "Visual Count"], ascending=[False, False]),
            use_container_width=True,
            hide_index=True,
        )
        st.caption("Prioritise rows where **Used In Report = Yes**. Those errors affect active pages, not just unused model objects.")

# ── TAB 2: Report usage ─────────────────────────────────────────────────────
with tab_report:
    missing_rows = analysis["report_missing_rows"]
    c1, c2 = st.columns(2)
    with c1:
        st.metric("Fields Referenced On Pages", len(analysis["usage_rows"]))
    with c2:
        st.metric("Referenced But Missing", len(missing_rows))

    if missing_rows:
        st.markdown("#### Visuals Using Fields Missing From The Model")
        st.dataframe(pd.DataFrame(missing_rows), use_container_width=True, hide_index=True)
    else:
        st.success("✅ Every field referenced by the report pages still exists in the model.")

    with st.expander("All fields used on report pages", expanded=False):
        st.dataframe(pd.DataFrame(analysis["usage_rows"]), use_container_width=True, hide_index=True)

# ── TAB 3: Naming ─────────────────────────────────────────────────────────────
with tab_name:
    issues = analysis["naming_issues"]
    s1, s2 = st.columns([1, 3])
    with s1:
        st.metric("Compliance", f"{analysis['naming_score']}%")
        st.progress(analysis["naming_score"]/100)
    with s2:
        st.markdown(
            "**Convention enforced:** PascalCase, no spaces, no underscores, "
            "no special characters, avoid abbreviations.")
        st.markdown(f"**{len(issues)} issues** across {analysis['measure_count']} measures.")

    if not issues:
        st.success("✅ All measures follow naming conventions!")
    else:
        show_fixed = st.checkbox("Show fixed", key="nm_show_fixed")
        for i, iss in enumerate(issues):
            if iss["fixed"] and not show_fixed:
                continue
            status = "✅ Fixed" if iss["fixed"] else ("❌" if iss["severity"] == "High" else "⚠️")
            with st.expander(f"{status} [{iss['table']}] **{iss['measure']}** → {iss['suggestion']}",
                             expanded=False):
                ca, cb, cc = st.columns([2, 2, 1])
                with ca:
                    st.markdown("**🔍 Problem**")
                    st.markdown(f"`{iss['measure']}` — {iss['reason']}")
                    st.markdown(f"_Severity:_ {_priority_badge('High' if iss['severity']=='High' else 'Medium')}",
                                unsafe_allow_html=True)
                with cb:
                    st.markdown("**🔧 Exact Fix**")
                    st.code(f"Rename:\n  {iss['measure']}\n→ {iss['suggestion']}", language="text")
                with cc:
                    st.markdown("**📌 Impact**")
                    # Real, ground-truth rename impact computed from DAX dependency graph
                    n_imp = iss.get("impact_total", 0)
                    if n_imp == 0:
                        st.markdown(
                            "<span class='impact-chip' style='background:#dcfce7;color:#166534;'>"
                            "✅ Safe — no dependents</span>", unsafe_allow_html=True)
                    elif n_imp <= 3:
                        st.markdown(
                            f"<span class='impact-chip' style='background:#fef3c7;color:#92400e;'>"
                            f"⚠ Breaks {n_imp} measure(s)</span>", unsafe_allow_html=True)
                    else:
                        st.markdown(
                            f"<span class='impact-chip' style='background:#fee2e2;color:#991b1b;'>"
                            f"🔴 Breaks {n_imp} measures</span>", unsafe_allow_html=True)
                    if iss.get("impact_examples"):
                        st.caption("e.g. " + ", ".join(f"`{x}`" for x in iss["impact_examples"]))

                    if not iss["fixed"]:
                        if st.button("Stage Rename", key=f"fix_nm_{i}", type="primary"):
                            _stage("renames", {
                                "table":        iss["table"],
                                "current_name": iss["measure"],
                                "new_name":     iss["suggestion"],
                            })
                            iss["fixed"] = True
                            analysis["naming_score"] = min(100, analysis["naming_score"] + int(60/max(len(issues),1)))
                            st.session_state[_fp] = analysis
                            st.rerun()
                st.caption("💡 Why this matters: PascalCase is the Power BI community standard. "
                           "Consistent names make every new developer productive immediately.")

        st.markdown("#### 📋 Export")
        df = pd.DataFrame([{
            "Table": i["table"], "Current": i["measure"], "Suggested": i["suggestion"],
            "Reason": i["reason"], "Severity": i["severity"],
            "Status": "Fixed" if i["fixed"] else "Open",
        } for i in issues])
        st.dataframe(df, use_container_width=True, hide_index=True)

# ── TAB 4: Descriptions ──────────────────────────────────────────────────────
with tab_desc:
    desc = analysis["desc_issues"]
    missing = [d for d in desc if not d["has_desc"]]
    s1, s2 = st.columns([1, 3])
    with s1:
        st.metric("Coverage", f"{analysis['desc_score']}%")
        st.progress(analysis["desc_score"]/100)
        st.markdown(
            f"<div style='background:#fef2f2;border-radius:8px;padding:10px 14px;margin-top:6px;'>"
            f"<strong style='color:#b91c1c;'>❌ {len(missing)} missing</strong><br/>"
            f"<span style='font-size:.82rem;color:#555;'>out of {len(desc)} measures</span></div>",
            unsafe_allow_html=True)
    with s2:
        st.markdown("""**Why this matters:**
Measures without descriptions force users to guess what a number means. Descriptions appear as
**tooltips in Power BI Desktop's field list**, in documentation tools, and in AI-assisted reporting.
Teams with full coverage onboard new members **40% faster** and reduce duplicate-measure incidents by ~60%.""")

    flt = st.selectbox("Filter", ["Missing only", "All", "Present only"], key="desc_flt")
    for i, d in enumerate(desc):
        show = (flt == "All") or (flt == "Missing only" and not d["has_desc"]) or \
               (flt == "Present only" and d["has_desc"])
        if not show: continue
        label = (f"✅ {d['name']} [{d['table']}]"
                 if d["has_desc"] else
                 f"❌ {d['name']} [{d['table']}] — NO DESCRIPTION")
        with st.expander(label, expanded=False):
            if d["has_desc"]:
                st.markdown(f"<div style='background:#f0fdf4;border-left:3px solid #4ade80;"
                            f"padding:10px 14px;border-radius:5px;font-size:.88rem;'>"
                            f"📝 {d['description']}</div>", unsafe_allow_html=True)
            else:
                cl, cr = st.columns([3, 1])
                with cl:
                    st.markdown("**🤖 AI-Generated Description (ready to copy):**")
                    st.markdown(f"<div style='background:#eff6ff;border-left:3px solid #60a5fa;"
                                f"padding:12px 16px;border-radius:6px;font-size:.88rem;color:#1e3a8a;'>"
                                f"💡 {d['ai_suggest']}</div>", unsafe_allow_html=True)
                    st.caption("→ Copy into Power BI Desktop → select measure → Properties → Description")
                with cr:
                    if st.button("Stage Description", key=f"fix_d_{i}", type="primary"):
                        _stage("descriptions", {
                            "table":       d["table"],
                            "measure":     d["name"],
                            "description": d["ai_suggest"],
                        })
                        d["description"] = d["ai_suggest"]; d["has_desc"] = True
                        new = sum(1 for x in desc if x["has_desc"])
                        analysis["desc_score"] = int(new/max(len(desc),1)*100)
                        st.session_state[_fp] = analysis
                        st.rerun()

# ── TAB 3: Placement ─────────────────────────────────────────────────────────
with tab_place:
    import collections as _col
    mp = analysis["misplaced"]

    # ── Cross-reference with report usage ─────────────────────────────────
    usage_idx = {row["Object"]: row for row in analysis.get("usage_rows", [])}

    def _placement_priority(m: dict) -> str:
        key = f"{m['current_table']}[{m['name']}]"
        used = bool(usage_idx.get(key))
        has_dependents = m.get("dependents", 0) > 0
        if used and has_dependents:
            return "Critical"
        if used or has_dependents:
            return "High"
        return "Low"

    # Detect existing measure tables from the loaded model
    existing_mt = sorted({
        t["name"] for t in meta.get("tables", [])
        if not t.get("is_internal")
        and t.get("measures")
        and all(c.get("is_hidden", False) for c in t.get("columns", []))
    })

    # ── Header ─────────────────────────────────────────────────────────────
    h1, h2, h3, h4 = st.columns(4)
    with h1:
        st.metric("Placement Score", f"{analysis['placement_score']}%")
        st.progress(analysis["placement_score"] / 100)
    with h2:
        st.metric("Misplaced Measures", len(mp))
    with h3:
        source_tables = len({m["current_table"] for m in mp})
        st.metric("Source Tables Affected", source_tables)
    with h4:
        st.metric("Existing Measure Tables",
                  len(existing_mt) if existing_mt else "None — create one")

    st.markdown("---")

    if not mp:
        st.success("✅ All measures are correctly placed in dedicated measure tables!")
        st.markdown("Your model follows the **`_Measures`** pattern. Measures are safe from "
                    "accidental deletion and the field list is clean for report authors.")
    else:
        # ── Situation banner ───────────────────────────────────────────────
        if existing_mt:
            st.info(
                f"**You already have {len(existing_mt)} dedicated measure table(s): "
                f"`{'`, `'.join(existing_mt)}`** — "
                "the {len(mp)} measures below just need to be moved into them. "
                "No setup required."
            )
        else:
            st.warning(
                "**No dedicated measure table found.** You need to create one in "
                "Power BI Desktop first (see instructions at the bottom of this tab), "
                "then move the measures below into it."
            )

        # ── Priority summary bar ───────────────────────────────────────────
        priorities = [_placement_priority(m) for m in mp]
        n_crit = priorities.count("Critical")
        n_high = priorities.count("High")
        n_low  = priorities.count("Low")
        p_cols = st.columns(3)
        with p_cols[0]:
            st.markdown(
                f"<div style='background:#fef2f2;border-left:4px solid #dc2626;border-radius:6px;"
                f"padding:10px 14px;'><b style='color:#991b1b'>🔴 Critical: {n_crit}</b><br/>"
                f"<span style='font-size:.78rem;color:#6b7280'>Used in report AND has dependents</span></div>",
                unsafe_allow_html=True)
        with p_cols[1]:
            st.markdown(
                f"<div style='background:#fff7ed;border-left:4px solid #ea580c;border-radius:6px;"
                f"padding:10px 14px;'><b style='color:#9a3412'>🟠 High: {n_high}</b><br/>"
                f"<span style='font-size:.78rem;color:#6b7280'>Used in report OR has dependents</span></div>",
                unsafe_allow_html=True)
        with p_cols[2]:
            st.markdown(
                f"<div style='background:#f0fdf4;border-left:4px solid #16a34a;border-radius:6px;"
                f"padding:10px 14px;'><b style='color:#166534'>🟢 Low: {n_low}</b><br/>"
                f"<span style='font-size:.78rem;color:#6b7280'>Not used in any active visual</span></div>",
                unsafe_allow_html=True)

        st.markdown("---")

        # ── Grouped by source table ────────────────────────────────────────
        by_table: dict = _col.defaultdict(list)
        for m in mp:
            by_table[m["current_table"]].append(m)

        st.markdown("#### Measures to Move — Grouped by Source Table")
        st.caption(
            "Sorted by impact: tables whose measures are actively used in reports come first. "
            "Fix **Critical** and **High** priority rows before **Low** ones."
        )

        # Sort groups: most critical first
        def _group_priority_score(items):
            return sum(2 if _placement_priority(m) == "Critical"
                       else 1 if _placement_priority(m) == "High" else 0
                       for m in items)

        sorted_groups = sorted(by_table.items(),
                               key=lambda kv: -_group_priority_score(kv[1]))

        for src_table, measures_list in sorted_groups:
            target = measures_list[0]["suggested_table"]
            group_score = _group_priority_score(measures_list)
            group_icon = "🔴" if group_score >= len(measures_list) * 2 else \
                         "🟠" if group_score > 0 else "🟢"

            with st.expander(
                f"{group_icon} **{src_table}** — {len(measures_list)} measure(s) → move to `{target}`",
                expanded=(group_score > 0)
            ):
                for i, m in enumerate(sorted(measures_list,
                                             key=lambda x: (
                                                 0 if _placement_priority(x) == "Critical"
                                                 else 1 if _placement_priority(x) == "High" else 2
                                             ))):
                    key_obj  = f"{m['current_table']}[{m['name']}]"
                    usage    = usage_idx.get(key_obj, {})
                    used_in  = bool(usage)
                    pages    = usage.get("Pages", "-") if used_in else "-"
                    vis_cnt  = usage.get("Visual Count", 0)
                    deps     = m.get("dependents", 0)
                    priority = _placement_priority(m)
                    pcolor   = {"Critical": "#dc2626", "High": "#ea580c", "Low": "#16a34a"}[priority]
                    pbg      = {"Critical": "#fef2f2", "High": "#fff7ed", "Low": "#f0fdf4"}[priority]

                    st.markdown(
                        f"<div style='border:1px solid #e5e7eb;border-radius:8px;"
                        f"padding:12px 16px;margin-bottom:10px;background:#fafafa;'>"
                        f"<div style='display:flex;align-items:center;gap:10px;flex-wrap:wrap;'>"
                        f"<b style='font-size:.95rem'>{m['name']}</b>"
                        f"<span style='background:{pbg};color:{pcolor};border-radius:4px;"
                        f"padding:2px 8px;font-size:.72rem;font-weight:700;'>{priority}</span>"
                        + (f"<span style='background:#dbeafe;color:#1d4ed8;border-radius:4px;"
                           f"padding:2px 8px;font-size:.72rem;'>📊 {vis_cnt} visual(s) · {pages}</span>"
                           if used_in else
                           "<span style='background:#f3f4f6;color:#6b7280;border-radius:4px;"
                           "padding:2px 8px;font-size:.72rem;'>Not on any page</span>")
                        + (f"<span style='background:#ede9fe;color:#6d28d9;border-radius:4px;"
                           f"padding:2px 8px;font-size:.72rem;'>⛓ {deps} dependent(s)</span>"
                           if deps else "")
                        + f"</div>"
                        + (f"<div style='margin-top:6px;font-size:.8rem;color:#6b7280;"
                           f"font-family:monospace;background:#f1f5f9;border-radius:4px;"
                           f"padding:4px 8px;'>{m.get('expression','')}</div>"
                           if m.get("expression") else "")
                        + f"</div>",
                        unsafe_allow_html=True,
                    )

                # Per-group fix instructions
                st.markdown(
                    f"<div style='background:#eff6ff;border-left:4px solid #3b82f6;"
                    f"border-radius:6px;padding:10px 14px;margin-top:8px;font-size:.84rem;'>"
                    f"<b>🔧 Steps for `{src_table}` → `{target}`:</b><br/>"
                    f"In Power BI Desktop → Model view: right-click each measure below in "
                    f"<code>{src_table}</code>, choose <b>Cut</b>, then right-click "
                    f"<code>{target}</code> and choose <b>Paste</b>.<br/>"
                    f"Measures to move: "
                    + ", ".join(f"<code>{m['name']}</code>" for m in measures_list)
                    + "</div>",
                    unsafe_allow_html=True,
                )

                # Stage-all button
                all_fixed = all(m["fixed"] for m in measures_list)
                if not all_fixed:
                    if st.button(f"Stage All {len(measures_list)} moves from `{src_table}`",
                                 key=f"fix_pg_{src_table}", use_container_width=True):
                        for m in measures_list:
                            if not m["fixed"]:
                                _stage("placements", {
                                    "measure":       m["name"],
                                    "current_table": m["current_table"],
                                    "target_table":  m["suggested_table"],
                                })
                                m["fixed"] = True
                        fc = sum(1 for x in mp if x["fixed"])
                        analysis["placement_score"] = int(fc / max(len(mp), 1) * 100) if mp else 100
                        st.session_state[_fp] = analysis
                        st.rerun()
                else:
                    st.success(f"✅ All moves from `{src_table}` staged.")

        st.markdown("---")

        # ── How to create a _Measures table (only if none exists) ──────────
        if not existing_mt:
            with st.expander("📖 How to create a dedicated measure table", expanded=False):
                st.markdown("""
**Option 1 — Enter Data (recommended, fastest)**
1. In Power BI Desktop → **Home** ribbon → **Enter Data**
2. Name the table `_Measures` (the underscore keeps it at the top of the field list)
3. Delete the default `Column1` column header, leave the table empty
4. Click **Load**
5. In Model view, hide the placeholder column (if any) so it doesn't appear in the field list

**Option 2 — DAX calculated table**
In Power BI Desktop → **Modeling** → **New Table**:
```dax
_Measures = ROW("x", BLANK())
```
Then hide the `x` column.

Once the table exists, right-click each measure in its current fact/dim table → **Cut** → right-click `_Measures` → **Paste**.
""")

        # ── Bulk export table ──────────────────────────────────────────────
        with st.expander("📋 All misplaced measures — full list", expanded=False):
            bulk_rows = []
            for m in mp:
                key_obj = f"{m['current_table']}[{m['name']}]"
                usage   = usage_idx.get(key_obj, {})
                bulk_rows.append({
                    "Priority":       _placement_priority(m),
                    "Measure":        m["name"],
                    "Current Table":  m["current_table"],
                    "Target Table":   m["suggested_table"],
                    "Used In Report": "Yes" if usage else "No",
                    "Visual Count":   usage.get("Visual Count", 0),
                    "Dependents":     m.get("dependents", 0),
                    "Expression":     m.get("expression", ""),
                    "Staged":         "Yes" if m["fixed"] else "No",
                })
            st.dataframe(pd.DataFrame(bulk_rows).sort_values(
                by=["Priority", "Visual Count"],
                key=lambda s: s.map({"Critical": 0, "High": 1, "Low": 2}).fillna(3)
                              if s.name == "Priority" else -s
            ), use_container_width=True, hide_index=True)


# ── TAB 6: Folders ───────────────────────────────────────────────────────────
with tab_fold:
    nf = analysis["no_folder"]
    s1, s2 = st.columns([1, 3])
    with s1:
        st.metric("Folder Score", f"{analysis['folder_score']}%")
        st.progress(analysis["folder_score"]/100)
    with s2:
        st.markdown("""**Why folders matter:** Without folders the field list is a flat unsorted list.
Report authors spend **3× longer** finding the right measure. Folders group metrics by domain
(Revenue · Costs · KPIs · YOY) making the model navigable by non-technical users.""")

    if not nf:
        st.success("✅ All measures have display folders!")
    else:
        st.caption("Suggested folder assignment based on measure names. Apply each via "
                   "Power BI Desktop → Measure Properties → Display Folder.")
        suggested = {
            "Revenue":          [m for m in nf if any(k in m["name"].lower() for k in ("sale","revenue","income","order","total"))],
            "Costs":            [m for m in nf if any(k in m["name"].lower() for k in ("cost","cogs","expense"))],
            "Margins & KPIs":   [m for m in nf if any(k in m["name"].lower() for k in ("margin","pct","rate","kpi","growth","yoy"))],
            "Budget":           [m for m in nf if any(k in m["name"].lower() for k in ("budget","variance","forecast"))],
            "Customers":        [m for m in nf if any(k in m["name"].lower() for k in ("customer","client","active","lifetime"))],
        }
        seen = set()
        for ms in suggested.values():
            for m in ms: seen.add(m["name"])
        suggested["Other"] = [m for m in nf if m["name"] not in seen]

        for folder, ms in suggested.items():
            if not ms: continue
            with st.expander(f"📁 {folder}  ({len(ms)} measures)", expanded=False):
                for j, m in enumerate(ms):
                    icon = "✅" if m["fixed"] else "📊"
                    cA, cB = st.columns([4, 1])
                    with cA:
                        st.markdown(
                            f"<div style='background:#faf5ff;border:1px solid #e9d5ff;"
                            f"border-radius:5px;padding:7px 12px;margin-bottom:4px;font-size:.86rem;'>"
                            f"{icon} <b>{m['name']}</b> "
                            f"<span style='color:#9ca3af;font-size:.78rem;'>[{m['table']}]</span> "
                            f"&nbsp;→ <i>{folder}</i></div>",
                            unsafe_allow_html=True)
                    with cB:
                        if not m["fixed"]:
                            if st.button("Stage Assign", key=f"fix_f_{folder}_{j}"):
                                _stage("display_folders", {
                                    "table":   m["table"],
                                    "measure": m["name"],
                                    "folder":  folder,
                                })
                                m["fixed"] = True
                                fc = sum(1 for x in nf if x["fixed"])
                                analysis["folder_score"] = int(fc/max(len(nf),1)*100) if nf else 100
                                st.session_state[_fp] = analysis
                                st.rerun()

# ── TAB 7: Bidirectional Relationships ───────────────────────────────────────
with tab_rel:
    bidir = analysis["bidir_rels"]
    if not bidir:
        st.success("✅ No bidirectional relationships detected — your model is clean!")
        st.markdown("Bidirectional cross-filters are the leading cause of incorrect totals in "
                    "Power BI. Your model avoids them entirely — that's excellent.")
    else:
        st.error(f"❌ {len(bidir)} bidirectional relationship(s) — this is **Critical**.")
        for i, r in enumerate(bidir):
            with st.expander(f"⚠️ {r['from_table']} ↔ {r['to_table']}", expanded=False):
                ca, cb = st.columns(2)
                with ca:
                    st.markdown("**🔍 Problem**")
                    st.markdown(
                        f"Relationship `{r['from_table']}[{r.get('from_column','')}]` ↔ "
                        f"`{r['to_table']}[{r.get('to_column','')}]` filters in **both directions**.")
                    st.markdown("""**📌 Why Critical:**
Bidirectional filters cause slicers to affect tables they shouldn't. Totals change unexpectedly
when filters are applied — leading to **wrong business decisions** based on wrong numbers.""")
                with cb:
                    st.markdown("**🔧 Exact Fix**")
                    st.code(
                        f"1. Open Model view in Power BI Desktop\n"
                        f"2. Click the relationship line between\n"
                        f"   {r['from_table']} → {r['to_table']}\n"
                        f"3. In Properties pane:\n"
                        f"   Cross-filter direction: Both → Single\n"
                        f"4. Save and verify all totals across visuals",
                        language="text")
                    st.markdown("**Outcome:** ✅ Filters flow in one direction — totals become predictable.")

# ── TAB 8: Accuracy (DAX-aware ground truth) ─────────────────────────────────
with tab_acc:
    broken  = analysis["broken_refs"]
    smells  = analysis["dax_smells"]
    orphans = analysis["orphans"]

    s1, s2 = st.columns([1, 3])
    with s1:
        st.metric("Accuracy Score", f"{analysis['accuracy_score']}%",
                  help="100% = no broken DAX references. Each broken reference deducts 10 points.")
        st.progress(analysis["accuracy_score"]/100)
        st.caption(
            f"🤖 {analysis['hidden_count']} hidden · "
            f"{analysis['time_intel_count']} time-intel measures detected — "
            "they are excluded from naming/placement/folder warnings to reduce false positives."
        )
    with s2:
        st.markdown("""**This tab is computed by parsing your actual DAX expressions** — not by
guessing from names. The findings below are the **highest-confidence** issues
in your model. Fix these *first*; they represent real bugs and dead code, not style preferences.""")

    st.markdown("---")

    # ── 1. BROKEN REFERENCES ─────────────────────────────────────────────────
    st.markdown(f"#### 🔴 Broken DAX references ({len(broken)})")
    st.caption(
        "Measures whose DAX references a measure / column / table that does not exist. "
        "These will fail at runtime or have already been silently failing. **100% confidence.**"
    )
    if not broken:
        st.success("✅ All DAX references resolve to existing model objects.")
    else:
        df_broken = pd.DataFrame([{
            "Measure":    b["measure"],
            "Kind":       b["kind"],
            "Missing":    b["missing"],
            "Detail":     b["message"],
        } for b in broken])
        st.dataframe(df_broken, use_container_width=True, hide_index=True)
        st.markdown(
            "**👉 Likely causes:** typos in measure / table names, items renamed without "
            "updating dependents, or columns deleted while measures still reference them. "
            "Use the rename-impact preview in the **Naming** tab before further renames."
        )

    st.markdown("---")

    # ── 2. DAX SMELLS ────────────────────────────────────────────────────────
    st.markdown(f"#### ⚡ DAX performance / style smells ({len(smells)})")
    st.caption(
        "Patterns that work, but are suboptimal. Severity reflects performance and "
        "readability impact, not correctness."
    )
    if not smells:
        st.success("✅ No common DAX anti-patterns detected.")
    else:
        sev_color = {"high":"#dc2626","medium":"#d97706","low":"#0891b2"}
        for sm in smells:
            color = sev_color.get(sm["severity"], "#6b7280")
            st.markdown(
                f"<div style='border-left:4px solid {color};background:#fafafa;"
                f"padding:10px 14px;border-radius:4px;margin-bottom:8px;'>"
                f"<span style='color:{color};font-weight:600;text-transform:uppercase;"
                f"font-size:.7rem;letter-spacing:.5px;'>{sm['severity']}</span> "
                f"<code>[{sm.get('table','')}]</code> "
                f"<strong>{sm['measure']}</strong><br/>"
                f"<span style='font-size:.88rem;color:#374151;'>{sm['message']}</span>"
                f"</div>", unsafe_allow_html=True)

    st.markdown("---")

    # ── 3. ORPHAN MEASURES ───────────────────────────────────────────────────
    st.markdown(f"#### 👻 Orphan measures ({len(orphans)})")
    st.caption(
        "Measures that no other measure references. Some are top-level KPIs displayed in "
        "reports — that's fine. But many are forgotten experiments that bloat your model. "
        "Review each: keep, hide, or delete."
    )
    if not orphans:
        st.success("✅ Every measure is referenced by at least one other measure.")
    else:
        # show in chunks of 4 columns
        cols_per_row = 4
        for chunk_start in range(0, min(len(orphans), 60), cols_per_row):
            cols = st.columns(cols_per_row)
            for i, name in enumerate(orphans[chunk_start:chunk_start + cols_per_row]):
                with cols[i]:
                    st.markdown(
                        f"<div style='background:#f9fafb;border:1px solid #e5e7eb;"
                        f"border-radius:5px;padding:6px 10px;margin:3px 0;font-size:.8rem;'>"
                        f"👻 {name}</div>", unsafe_allow_html=True)
        if len(orphans) > 60:
            st.caption(f"…and {len(orphans) - 60} more.")

# ── TAB 9: Security / RLS ────────────────────────────────────────────────────
with tab_sec:
    sec = analysis["security_audit"]
    findings = sec["findings"]

    s1, s2 = st.columns([1, 3])
    with s1:
        st.metric("Security Score", f"{analysis['security_score']}%")
        st.progress(analysis["security_score"] / 100)
        st.markdown(
            f"<div style='background:#fdf2f8;border-radius:8px;padding:10px 14px;margin-top:6px;'>"
            f"<strong style='color:#9d174d;'>🔐 {len(sec['roles'])} RLS role(s)</strong><br/>"
            f"<span style='font-size:.82rem;color:#555;'>{len(findings)} finding(s)</span></div>",
            unsafe_allow_html=True,
        )
    with s2:
        st.markdown("""**Why this matters:** Governance without security is incomplete.
Power BI datasets often hold customer, payroll, sales, or region-level data that must be restricted.
If RLS is missing or weak, the model is operationally risky even if the DAX is perfect.""")

    if sec["sensitive_tables"]:
        st.caption("Sensitive-looking tables detected: " + ", ".join(f"`{t}`" for t in sec["sensitive_tables"][:8]))

    if not findings:
        st.success("✅ No major RLS governance issues detected.")
    else:
        for idx, f in enumerate(findings):
            icon = "🔴" if f["severity"] == "Critical" else ("🟠" if f["severity"] == "High" else "🟡")
            with st.expander(f"{icon} {f['severity']} — {f['finding']}", expanded=False):
                st.markdown("**Detail**")
                st.markdown(f["detail"])
                st.markdown("**Recommended Fix**")
                st.code(f["fix"], language="text")

# ── TAB 10: Schema Health ────────────────────────────────────────────────────
with tab_schema:
    sch = analysis["schema_audit"]
    naming = sch["table_naming_issues"]
    isolated = sch["isolated_tables"]
    m2m = sch["m2m_rels"]
    metric_cols = sch["metric_calc_cols"]
    inactive = sch["inactive_rels"]

    s1, s2 = st.columns([1, 3])
    with s1:
        st.metric("Schema Score", f"{analysis['schema_score']}%")
        st.progress(analysis["schema_score"] / 100)
    with s2:
        st.markdown("""**What senior BI teams care about:** clear Fact/Dim naming, no ambiguous many-to-many joins,
no orphaned tables, and no aggregation logic trapped in calculated columns. These issues directly affect
model maintainability, memory size, and trust in report totals.""")

    st.markdown("#### Table Naming")
    if not naming:
        st.success("✅ All data tables follow a recognisable naming convention.")
    else:
        st.dataframe(pd.DataFrame([{
            "Table": i["table"],
            "Issue": i["reason"],
            "Suggested": i["suggestion"],
        } for i in naming]), use_container_width=True, hide_index=True)

    st.markdown("#### Relationship Design")
    if not isolated and not m2m and not inactive:
        st.success("✅ No isolated tables, many-to-many joins, or inactive relationships detected.")
    else:
        if isolated:
            st.markdown("**Isolated tables**")
            st.dataframe(pd.DataFrame([{"Table": t["name"], "Issue": "No relationships to any other table"} for t in isolated]),
                         use_container_width=True, hide_index=True)
        if m2m:
            st.markdown("**Many-to-many relationships**")
            st.dataframe(pd.DataFrame([{
                "From": r["from_table"],
                "To": r["to_table"],
                "Issue": "Many-to-many relationship",
                "Fix": "Introduce a bridge table or remodel keys",
            } for r in m2m]), use_container_width=True, hide_index=True)
        if inactive:
            st.markdown("**Inactive relationships**")
            st.dataframe(pd.DataFrame([{
                "From": r["from_table"],
                "To": r["to_table"],
                "Issue": "Inactive relationship",
            } for r in inactive]), use_container_width=True, hide_index=True)

    st.markdown("#### Calculated Columns That Should Probably Be Measures")
    if not metric_cols:
        st.success("✅ No aggregation-style calculated columns detected.")
    else:
        st.dataframe(pd.DataFrame([{
            "Table": c.get("table", ""),
            "Column": c["name"],
            "Expression": c.get("expression", "")[:120],
            "Recommendation": "Convert to measure",
        } for c in metric_cols]), use_container_width=True, hide_index=True)

# ─────────────────────────────────────────────────────────────────────────────
# DELIVER REAL FIXES — patched .pbit + Tabular Editor rename script
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown("### 📦 Apply Fixes to Your Real Report")

s = _staged()
n_desc   = len(s["descriptions"])
n_fold   = len(s["display_folders"])
n_place  = len(s["placements"])
n_rename = len(s["renames"])
total    = n_desc + n_fold + n_place + n_rename

if total == 0:
    st.info(
        "No fixes staged yet. Use the issue-level actions above, such as **Stage Move**, "
        "**Stage Description**, **Stage Assign**, or **Stage Rename**, to prepare fixes for export."
    )
else:
    st.markdown(
        f"<div style='background:#f0fdf4;border:1px solid #86efac;border-radius:10px;"
        f"padding:14px 18px;margin-bottom:14px;'>"
        f"<strong style='color:#166534;'>📋 {total} change(s) staged:</strong> "
        f"{n_desc} descriptions · {n_fold} display-folders · {n_place} measure moves · "
        f"<strong>{n_rename} renames</strong> (script-only)"
        f"</div>",
        unsafe_allow_html=True,
    )

    # ── Two-column layout: safe rewrites (LEFT)  +  rename script (RIGHT) ──
    dl1, dl2 = st.columns(2, gap="large")

    # ── LEFT: Patched .pbit (Option B) ──────────────────────────────────────
    with dl1:
        st.markdown("#### 🟢 Safe Auto-Fix — Patched .pbit")
        st.caption(
            "Descriptions, display folders, and measure moves are applied directly "
            "to the model. **Renames are NOT included** here because they would "
            "break visual bindings — use the Tabular Editor script for those."
        )

        if not raw_bytes:
            st.warning("Upload a real .pbit / .pbix above to enable patched downloads (demo data has no source file).")
        elif n_desc + n_fold + n_place == 0:
            st.info("No safe fixes staged — stage descriptions, folders, or moves first.")
        else:
            try:
                patched_bytes, counters = apply_safe_fixes(
                    raw_bytes,
                    descriptions=s["descriptions"],
                    display_folders=s["display_folders"],
                    placements=s["placements"],
                )
                base = (file_label.rsplit(".", 1)[0] if "." in file_label else file_label) + "_fixed"
                ext  = "pbit" if file_label.lower().endswith(".pbit") else "pbix"
                st.download_button(
                    label=f"⬇ Download {base}.{ext}",
                    data=patched_bytes,
                    file_name=f"{base}.{ext}",
                    mime="application/octet-stream",
                    use_container_width=True,
                    type="primary",
                )
                st.caption(
                    f"Applied: {counters['descriptions_applied']} descriptions · "
                    f"{counters['folders_applied']} folders · "
                    f"{counters['measures_moved']} moves"
                    + (" · created _Measures table" if counters["measures_table_created"] else "")
                    + (f" · {counters['descriptions_skipped'] + counters['folders_skipped'] + counters['moves_skipped']} skipped"
                       if any(counters[k] for k in ("descriptions_skipped","folders_skipped","moves_skipped")) else "")
                )
                st.markdown(
                    "**Next steps:** Open the downloaded file in Power BI Desktop. "
                    "Verify the changes look correct, then save."
                )
            except Exception as exc:
                st.error(f"Could not patch file: {exc}")

    # ── RIGHT: Tabular Editor rename script (Option C) ─────────────────────
    with dl2:
        st.markdown("#### 🔵 Safe Renames — Tabular Editor Script")
        st.caption(
            "Renames break visual bindings if not done carefully. Tabular Editor "
            "(free at tabulareditor.com) updates every dependent reference safely. "
            "Run the generated script against your model."
        )
        if n_rename == 0:
            st.info("No renames staged — stage from the Naming tab.")
        else:
            script_text = generate_rename_script(s["renames"])
            st.download_button(
                label=f"⬇ Download rename_script.csx ({n_rename} renames)",
                data=script_text.encode("utf-8"),
                file_name="rename_script.csx",
                mime="text/plain",
                use_container_width=True,
                type="primary",
            )
            with st.expander("👀 Preview script"):
                st.code(script_text, language="csharp")
            st.markdown(
                "**Steps:**\n"
                "1. Open `.pbix` in Power BI Desktop\n"
                "2. **External Tools → Tabular Editor**\n"
                "3. **File → Open Script…** → pick `rename_script.csx`\n"
                "4. Press **F5** to run\n"
                "5. **Ctrl+S** writes back to Power BI"
            )

    # ── Clear staged fixes ──────────────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    if st.button("🗑 Clear all staged fixes", type="secondary"):
        st.session_state["staged_fixes"] = {
            "descriptions": [], "display_folders": [],
            "placements":   [], "renames":         [],
        }
        st.rerun()

# ─────────────────────────────────────────────────────────────────────────────
# BUSINESS BOTTOM LINE
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("---")
broken_n = len(analysis["broken_measure_rows"])
broken_used_n = sum(1 for row in analysis["broken_measure_rows"] if row["Used In Report"] == "Yes")
report_missing_n = len(analysis["report_missing_rows"])
mp_n = len(analysis["misplaced"])
sec_n = len(analysis["security_audit"]["findings"])
st.markdown(f"""<div style="background:#f0f4ff;border:1px solid #c7d2fe;border-radius:12px;padding:18px 24px;">
<p style="margin:0;color:#3730a3;font-size:.92rem;line-height:1.7;">
💼 <strong>Summary:</strong> Your model currently has <strong>{broken_n} broken measure error(s)</strong>,
<strong>{broken_used_n} of them used on report pages</strong>,
<strong>{report_missing_n} field reference(s) on pages that no longer exist in the model</strong>,
<strong>{mp_n} misplaced measures</strong>, and <strong>{sec_n} security finding(s)</strong>.
Fix the broken objects and broken report bindings first. The optional cleanup tabs are lower priority.
</p></div>""", unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)
report_bytes = _build_excel_report(analysis)
st.download_button(
    label="⬇ Download governance_report.xlsx",
    data=report_bytes,
    file_name="governance_report.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)
st.caption("Use this workbook to share findings with BI leads, model owners, and governance reviewers.")
